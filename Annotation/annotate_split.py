# pip install pandas openpyxl

"""
split.py

Purpose
-------
Convert an input CSV annotation file into multiple balanced Excel annotation files.

Input CSV columns expected:
    question,gold_answer,KG answer,taxonomy_label,confidence,source

Output:
    annotation_outputs/annotation_group_1.xlsx
    annotation_outputs/annotation_group_2.xlsx
    annotation_outputs/annotation_group_3.xlsx

Each output file:
    - removes the confidence column
    - keeps a stable case_id for merging later
    - adds dropdown column: label_correctness
    - adds three Yes/No columns for Wikidata structural causes:
        missing_edge
        missing_node
        missing_property_or_qualifier
    - adds free-text column: note
    - uses alternating row colors for better readability

Important
---------
CSV files cannot store dropdown menus.
Therefore, this script outputs .xlsx files instead of .csv files.
"""

import pandas as pd
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Dropdown options
# ---------------------------------------------------------------------

# Annotators use this column to verify whether the assigned taxonomy label
# is correct according to answer-level comparison.
TAXONOMY_LABEL_CORRECT_OPTIONS = [
    "Correct",
    "Incorrect",
    "Unsure"
]

# Annotators use these values for each Wikidata structural cause column.
# Each cause is now annotated independently, so multiple causes can be marked.
YES_NO_OPTIONS = [
    "Yes",
    "No"
]


# ---------------------------------------------------------------------
# Helper 1: clean column names
# ---------------------------------------------------------------------

def clean_column_names(df):
    """
    Clean CSV column names.

    Why this is needed:
    Sometimes CSV files contain invisible characters such as BOM '\\ufeff',
    or spaces around column names.

    Example:
        ' taxonomy_label ' -> 'taxonomy_label'
        '\\ufeffquestion' -> 'question'
    """

    df.columns = (
        df.columns
        .astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
    )

    return df


# ---------------------------------------------------------------------
# Helper 2: balanced splitting
# ---------------------------------------------------------------------

def split_balanced(df, n_groups=3, stratify_cols=None):
    """
    Split a dataframe into n balanced groups.

    Parameters
    ----------
    df : pandas.DataFrame
        Full dataframe to split.

    n_groups : int
        Number of output groups.

    stratify_cols : list[str] or None
        Columns used for approximate stratified splitting.

        Recommended:
            stratify_cols=["taxonomy_label"]

        Not recommended:
            stratify_cols=["source"]

        Why not use source?
            In your file, source may be unique or almost unique for every row.
            If we group by source directly, each source group has only one row.
            A bad splitting function may then put every row into group 1.

    Returns
    -------
    list[pandas.DataFrame]
        A list of n_groups dataframes.
    """

    df = df.reset_index(drop=True).copy()

    groups = [[] for _ in range(n_groups)]

    if stratify_cols:
        for col in stratify_cols:
            if col not in df.columns:
                raise ValueError(
                    f"Stratification column not found: {col}\n"
                    f"Available columns: {list(df.columns)}"
                )

        grouped = df.groupby(stratify_cols, sort=False, dropna=False)

        # Global pointer keeps distribution balanced across all labels.
        pointer = 0

        for _, sub_df in grouped:
            sub_df = sub_df.reset_index(drop=True)

            for _, row in sub_df.iterrows():
                group_id = pointer % n_groups
                groups[group_id].append(row)
                pointer += 1

    else:
        # Simple round-robin split without stratification.
        for i, (_, row) in enumerate(df.iterrows()):
            group_id = i % n_groups
            groups[group_id].append(row)

    result = []

    for group_rows in groups:
        if len(group_rows) == 0:
            result.append(pd.DataFrame(columns=df.columns))
        else:
            result.append(
                pd.DataFrame(group_rows, columns=df.columns).reset_index(drop=True)
            )

    return result


# ---------------------------------------------------------------------
# Helper 3: robust Excel header lookup
# ---------------------------------------------------------------------

def find_column_index(header, target_name):
    """
    Find a column index in an Excel header row.

    Returns
    -------
    int
        1-based Excel column index.

    This function is more robust than:
        header.index("label_correctness")

    because it cleans hidden BOM characters and surrounding spaces.
    """

    cleaned_header = [
        str(h).replace("\ufeff", "").strip() if h is not None else ""
        for h in header
    ]

    if target_name not in cleaned_header:
        raise ValueError(
            f"Column '{target_name}' not found.\n"
            f"Available columns are:\n{cleaned_header}"
        )

    return cleaned_header.index(target_name) + 1


# ---------------------------------------------------------------------
# Helper 4: add dropdown menus and formatting to Excel
# ---------------------------------------------------------------------

def add_dropdowns_and_formatting(xlsx_path):
    """
    Add dropdown menus and formatting to one Excel annotation file.

    Dropdown columns:
        label_correctness:
            Correct / Incorrect / Unsure

        missing_edge:
            Yes / No

        missing_node:
            Yes / No

        missing_property_or_qualifier:
            Yes / No

    Free-text column:
        note

    Why multiple Wikidata cause columns?
        A case may involve more than one structural KG issue.
        For example, Wikidata may miss both an entity and a connecting statement.
        Separate Yes/No columns allow annotators to mark all causes that apply.
    """

    wb = load_workbook(xlsx_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    max_row = ws.max_row
    max_col = ws.max_column

    # Find target annotation columns.
    label_correctness_col = find_column_index(header, "label_correctness")
    missing_edge_col = find_column_index(header, "missing_edge")
    missing_node_col = find_column_index(header, "missing_node")
    missing_property_col = find_column_index(header, "missing_property_or_qualifier")

    # Convert column numbers to Excel letters.
    label_correctness_col_letter = get_column_letter(label_correctness_col)
    missing_edge_col_letter = get_column_letter(missing_edge_col)
    missing_node_col_letter = get_column_letter(missing_node_col)
    missing_property_col_letter = get_column_letter(missing_property_col)

    # -----------------------------------------------------------------
    # Dropdown validation
    # -----------------------------------------------------------------

    # Dropdown for taxonomy label correctness.
    dv_label_correctness = DataValidation(
        type="list",
        formula1=f'"{",".join(TAXONOMY_LABEL_CORRECT_OPTIONS)}"',
        allow_blank=True
    )

    # Dropdown for each Wikidata structural cause column.
    dv_yes_no = DataValidation(
        type="list",
        formula1=f'"{",".join(YES_NO_OPTIONS)}"',
        allow_blank=True
    )

    ws.add_data_validation(dv_label_correctness)
    ws.add_data_validation(dv_yes_no)

    if max_row >= 2:
        # Apply label correctness dropdown.
        dv_label_correctness.add(
            f"{label_correctness_col_letter}2:"
            f"{label_correctness_col_letter}{max_row}"
        )

        # Apply Yes/No dropdowns to all three Wikidata cause columns.
        dv_yes_no.add(
            f"{missing_edge_col_letter}2:"
            f"{missing_edge_col_letter}{max_row}"
        )

        dv_yes_no.add(
            f"{missing_node_col_letter}2:"
            f"{missing_node_col_letter}{max_row}"
        )

        dv_yes_no.add(
            f"{missing_property_col_letter}2:"
            f"{missing_property_col_letter}{max_row}"
        )

    # -----------------------------------------------------------------
    # Formatting colors
    # -----------------------------------------------------------------

    header_fill = PatternFill("solid", fgColor="1F4E79")       # dark blue
    header_font = Font(bold=True, color="FFFFFF")              # white text

    row_fill_odd = PatternFill("solid", fgColor="FFFFFF")      # white
    row_fill_even = PatternFill("solid", fgColor="EAF3F8")     # very light blue

    annotation_fill_odd = PatternFill("solid", fgColor="FFF2CC")   # light yellow
    annotation_fill_even = PatternFill("solid", fgColor="FCE4B2")  # deeper yellow

    thin_side = Side(style="thin", color="D9D9D9")

    border = Border(
        top=thin_side,
        left=thin_side,
        right=thin_side,
        bottom=thin_side
    )

    # -----------------------------------------------------------------
    # Header formatting
    # -----------------------------------------------------------------

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    ws.row_dimensions[1].height = 32

    # Clean header for formatting lookup.
    cleaned_header = [
        str(h).replace("\ufeff", "").strip() if h is not None else ""
        for h in header
    ]

    # These are the columns annotators fill in.
    annotation_col_names = [
        "label_correctness",
        "missing_edge",
        "missing_node",
        "missing_property_or_qualifier",
        "note"
    ]

    annotation_col_indices = []

    for col_name in annotation_col_names:
        if col_name in cleaned_header:
            annotation_col_indices.append(cleaned_header.index(col_name) + 1)

    # -----------------------------------------------------------------
    # Body formatting with alternating row colors
    # -----------------------------------------------------------------

    for row_idx in range(2, max_row + 1):
        is_even_body_row = row_idx % 2 == 0

        normal_fill = row_fill_even if is_even_body_row else row_fill_odd
        annotation_fill = annotation_fill_even if is_even_body_row else annotation_fill_odd

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Annotation columns remain yellow, but still alternate slightly.
            if col_idx in annotation_col_indices:
                cell.fill = annotation_fill
            else:
                cell.fill = normal_fill

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )
            cell.border = border

    # -----------------------------------------------------------------
    # Highlight annotation headers
    # -----------------------------------------------------------------

    for col_idx in annotation_col_indices:
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill("solid", fgColor="FFD966")     # yellow header
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    # -----------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------

    widths = {
        "case_id": 10,
        "question": 45,
        "gold_answer": 35,
        "KG answer": 35,
        "taxonomy_label": 30,
        "source": 24,
        "label_correctness": 24,
        "missing_edge": 18,
        "missing_node": 18,
        "missing_property_or_qualifier": 30,
        "note": 45,
    }

    for i, col_name in enumerate(cleaned_header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 22)

    # -----------------------------------------------------------------
    # Row heights
    # -----------------------------------------------------------------

    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 65

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Enable Excel filter.
    ws.auto_filter.ref = ws.dimensions

    # Save changes.
    wb.save(xlsx_path)


# ---------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------

def generate_annotation_files(
    input_csv,
    output_dir="annotation_outputs",
    n_groups=3,
    stratify_cols=None
):
    """
    Generate balanced annotation Excel files from one input CSV.

    Parameters
    ----------
    input_csv : str
        Path to the input CSV.

    output_dir : str
        Folder where output Excel files will be saved.

    n_groups : int
        Number of annotation files to create.

    stratify_cols : list[str] or None
        Recommended:
            ["taxonomy_label"]

        This keeps taxonomy label distribution more balanced across groups.

        Use None if you only want simple round-robin balancing.
    """

    input_csv = Path(input_csv)
    output_dir = Path(output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    # Read CSV.
    #
    # header=0 means:
    #     The first row is the column title/header row.
    #     It is NOT treated as data.
    df = pd.read_csv(input_csv, header=0)

    # Clean column names.
    df = clean_column_names(df)

    # Expected input columns.
    expected_cols = [
        "question",
        "gold_answer",
        "KG answer",
        "taxonomy_label",
        "confidence",
        "source"
    ]

    # Validate input columns.
    missing = [col for col in expected_cols if col not in df.columns]

    if missing:
        raise ValueError(
            f"Missing expected columns: {missing}\n"
            f"Actual columns are: {list(df.columns)}"
        )

    # Remove confidence column because the annotation task should not show it.
    df = df.drop(columns=["confidence"])

    # Add stable case_id.
    #
    # This is useful later when merging annotations from different files.
    # It preserves the original row identity.
    if "case_id" not in df.columns:
        df.insert(0, "case_id", range(1, len(df) + 1))

    # -----------------------------------------------------------------
    # Add annotation columns before splitting
    # -----------------------------------------------------------------

    # This column checks whether the automatic taxonomy label is correct.
    df["label_correctness"] = ""

    # Wikidata structural cause columns.
    #
    # These replace the old single wikidata_cause column.
    # Annotators can mark multiple causes as Yes.
    #
    # Example:
    #     missing_edge = Yes
    #     missing_node = No
    #     missing_property_or_qualifier = Yes
    #
    # This means both a missing edge and a missing property/qualifier
    # are relevant for the same case.
    df["missing_edge"] = ""
    df["missing_node"] = ""
    df["missing_property_or_qualifier"] = ""

    # Free-text note column.
    df["note"] = ""

    # Final column order in output files.
    final_cols = [
        "case_id",
        "question",
        "gold_answer",
        "KG answer",
        "taxonomy_label",
        "source",
        "label_correctness",
        "missing_edge",
        "missing_node",
        "missing_property_or_qualifier",
        "note"
    ]

    df = df[final_cols]

    print("Final columns before splitting:")
    print(list(df.columns))

    # Split into balanced groups.
    splits = split_balanced(
        df,
        n_groups=n_groups,
        stratify_cols=stratify_cols
    )

    output_paths = []

    # Save each group to an Excel file.
    for i, split_df in enumerate(splits, start=1):
        out_path = output_dir / f"annotation_group_{i}.xlsx"

        print(f"\nSaving group {i}: {len(split_df)} rows")
        print("Columns:", list(split_df.columns))

        # Save dataframe to Excel.
        split_df.to_excel(out_path, index=False)

        # Add dropdowns and formatting.
        add_dropdowns_and_formatting(out_path)

        output_paths.append(out_path)

    # Print summary.
    print("\nFinished.")
    print(f"Total input data rows: {len(df)}")

    for i, split_df in enumerate(splits, start=1):
        print(f"group {i}: {len(split_df)} rows")

    print("\nOutput files:")

    for path in output_paths:
        print(path)

    return output_paths


# ---------------------------------------------------------------------
# Run script
# ---------------------------------------------------------------------

if __name__ == "__main__":
    generate_annotation_files(
        input_csv="annotation_cases_235B_main_inconsistent.csv",
        output_dir="annotation_outputs",
        n_groups=3,

        # Recommended for your task:
        # balance the distribution of taxonomy labels across the 3 files.
        #
        # Do NOT stratify by source because source may be unique per row.
        stratify_cols=["taxonomy_label"]

        # Alternative:
        # Use this if you want pure round-robin splitting:
        # stratify_cols=None
    )