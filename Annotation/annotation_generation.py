#!/usr/bin/env python3
"""
build_annotation_235B_interactive_balanced_inconsistent.py

Summary
-------
This script builds balanced annotation CSV files for inconsistent Qwen3-235B
cases using the same input files as the original merge_taxonomy_full.py
pipeline.

For the 235B model, the script:
1. Reads the Simple-heuristics taxonomy CSV and the LLM-as-a-judge CSV.
2. Drops Simple-heuristics rows whose taxonomy_label is different_unclassified.
3. Drops all rows whose normalized taxonomy_label is Same.
4. Preserves the complete original file_path in the public `source` column,
   for example SimpleQA/NQ_table_test_simple/00503.json.
5. Removes rows whose source/file_path appears in
   Output_GRASP/script/large_results_report.txt, meaning the KG answer has
   more than 10 result rows.
6. Merges Simple-heuristics and LLM-as-a-judge rows into one shared internal
   schema.
7. Removes duplicate question rows, keeping the first occurrence.
8. Scans whether the requested target size can be perfectly balanced across
   the 2x2 source/method buckets:
   - SimpleQA / Simple heuristics
   - SimpleQA / LLM-as-a-judge
   - ComplexQA / Simple heuristics
   - ComplexQA / LLM-as-a-judge
9. If perfect 2x2 balance is impossible, prints the maximum possible perfect
   balanced size and asks whether to use that smaller size.
10. If the user enters y, writes the maximum perfect-balanced set.
11. If the user enters n, writes the requested target size using source-first
    balancing:
    - SimpleQA and ComplexQA are balanced first.
    - Method balance is preferred.
    - If one method is short, the other method complements within the same
      source group.
12. Writes the next 10 inconsistent presentation cases:
    - SimpleQA = 5
    - ComplexQA = 5
    - Method balance is preferred, but method shortages are complemented within
      the same source group.

The public output schema is exactly:

    question,gold_answer,KG answer,taxonomy_label,sparql,confidence,source

For Simple-heuristics rows, confidence is similarity_score.
For LLM-as-a-judge rows, confidence is difference_severity.

Run from the repository root with:

    python build_annotation_235B_interactive_balanced_inconsistent.py
"""

from __future__ import annotations

from pathlib import Path
import csv
import math
import re
from collections import Counter, defaultdict


# ---------------------------------------------------------------------------
# Public output schema for final annotation CSV files.
# The `source` column intentionally stores the complete dataset-relative JSON
# path, e.g. SimpleQA/NQ_table_test_simple/00503.json.
# ---------------------------------------------------------------------------
OUTPUT_COLUMNS = [
    "question",
    "gold_answer",
    "KG answer",
    "taxonomy_label",
    "sparql",
    "confidence",
    "source",
]


# Requested target size for the main annotation set.
# If this exact size cannot be perfectly 2x2 balanced, the script will ask
# whether to use the maximum possible perfect-balanced size instead.
TARGET_SELECTION_CASES = 300


# Size for the presentation example slice after the main selected set.
SECOND_SLICE_CASES = 10


# Report containing JSON files whose result row count is greater than 10.
LARGE_RESULTS_REPORT = Path("Output_GRASP/script/large_results_report.txt")


# Preferred order for LLM-as-a-judge difference severity.
# Lower rank means more confident and therefore earlier in the output.
DIFFERENCE_SEVERITY_ORDER = {
    "major": 0,
    "moderate": 1,
    "minor": 2,
    "none": 3,
}


# Public method names used internally for balancing.
SIMPLE_METHOD = "Simple heuristics"
LLM_METHOD = "LLM-as-a-judge"


# QA groups used for balancing.
SIMPLE_QA_GROUP = "SimpleQA"
COMPLEX_QA_GROUP = "ComplexQA"


# 235B-only input/output configuration.
MODEL_PAIR = {
    "label": "235B",
    "short_model": "Qwen3-235B",
    "model": "Qwen3-235B-Thinking",
    "judged_csv": Path("LLM_as_a_Judge/235B_judged_filtered.csv"),
    "taxonomy_csv": Path("Output_GRASP/Qwen3-235B-A22B-Thinking-2507-AWQ/all_valid_cases_with_taxonomy.csv"),
    "main_csv": Path("Annotation/annotation_cases_235B_main_inconsistent.csv"),
    "second10_csv": Path("Annotation/annotation_cases_235B_second10_inconsistent.csv"),
    "statistics_md": Path("Annotation/annotation_235B_interactive_inconsistent_statistics.md"),
}


def find_project_root() -> Path:
    """Locate the repository root from either the root or Output_GRASP/."""
    cwd = Path.cwd().resolve()
    judge_dir_names = ["LLM_as_a_Judge", "LLM-as-Judge"]

    for judge_dir_name in judge_dir_names:
        if (cwd / judge_dir_name).is_dir() and (cwd / "Output_GRASP").is_dir():
            return cwd

    if cwd.name == "Output_GRASP":
        parent = cwd.parent
        for judge_dir_name in judge_dir_names:
            if (parent / judge_dir_name).is_dir() and (parent / "Output_GRASP").is_dir():
                return parent

    raise RuntimeError(
        "Could not locate project root. Run from the repository root or "
        "from Output_GRASP/. Expected Output_GRASP plus one of: "
        "LLM_as_a_Judge, LLM-as-Judge."
    )


def normalize_taxonomy_label(label: str) -> str:
    """Normalize taxonomy labels so equivalent labels are counted together."""
    label = (label or "").strip()
    if label.lower() == "same":
        return "Same"
    return label


def is_inconsistent_taxonomy_label(label: str) -> bool:
    """Return True only for non-Same taxonomy labels."""
    return normalize_taxonomy_label(label) != "Same"


def should_keep_simple_heuristics_row(label: str) -> bool:
    """
    Exclude Simple-heuristics rows that were not assigned a usable class.

    This does not remove Same rows by itself. Same rows are removed by the
    inconsistent-case filter after normalization.
    """
    return (label or "").strip().lower() != "different_unclassified"


def normalize_file_path_key(file_path_value: str) -> str:
    """
    Normalize a source/file_path value for matching.

    Keeps the dataset-relative suffix beginning with SimpleQA/ or ComplexQA/.
    This makes absolute paths, relative paths, and report paths comparable.
    """
    if not file_path_value:
        return ""

    normalized = str(file_path_value).strip().replace("\\", "/")
    match = re.search(r"(?:^|/)((?:SimpleQA|ComplexQA)/[^\s]+\.json)$", normalized)
    if match:
        return match.group(1)
    return normalized


def qa_group_from_source(source: str) -> str:
    """Map a complete source path to SimpleQA, ComplexQA, or UNKNOWN."""
    key = normalize_file_path_key(source)
    if key.startswith("SimpleQA/"):
        return SIMPLE_QA_GROUP
    if key.startswith("ComplexQA/"):
        return COMPLEX_QA_GROUP
    return "UNKNOWN"


def normalize_difference_severity(value: str) -> str:
    """
    Normalize difference_severity values.

    The final value must be exactly one of:
    none, minor, moderate, major.
    """
    severity = (value or "").strip().lower().replace("_", " ").replace("-", " ")
    severity = re.sub(r"\s+", " ", severity)

    if severity in DIFFERENCE_SEVERITY_ORDER:
        return severity

    raise ValueError(
        f"Invalid difference_severity value: {value!r}. "
        "Expected exactly one of: none, minor, moderate, major."
    )


def parse_float(value: str) -> float:
    """Parse a floating-point value and return NaN if parsing fails."""
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return math.nan


def safe_sort_float(value: float) -> float:
    """Return a sortable float, sending missing values to the end."""
    if value is None or math.isnan(value):
        return float("-inf")
    return value


def parse_large_results_report(report_path: Path) -> dict[str, set[str]]:
    """
    Parse Output_GRASP/script/large_results_report.txt.

    Returns {model_output_directory_name: {dataset-relative JSON paths}}.
    """
    large_paths_by_model: dict[str, set[str]] = {}

    if not report_path.exists():
        print(f"WARNING: Large-results report not found: {report_path}")
        print("         No rows will be removed by long-answer filtering.")
        return large_paths_by_model

    line_pattern = re.compile(
        r"^(?P<model>\S+)\s+"
        r"(?P<count>\d+)\s+rows\s+"
        r".*?(?P<file_path>(?:SimpleQA|ComplexQA)/\S+\.json)\s*$"
    )

    with report_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            match = line_pattern.match(line.strip())
            if not match:
                continue
            model_dir = match.group("model")
            file_path_key = normalize_file_path_key(match.group("file_path"))
            large_paths_by_model.setdefault(model_dir, set()).add(file_path_key)

    return large_paths_by_model


def read_csv_rows(csv_path: Path) -> list[dict]:
    """Read a CSV file into a list of dictionaries."""
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def require_columns(csv_path: Path, rows: list[dict], required_columns: list[str]) -> bool:
    """Check whether a non-empty CSV contains all required columns."""
    if not rows:
        return True
    missing = [column for column in required_columns if column not in rows[0]]
    if missing:
        print(f"  Missing required columns in {csv_path}: {', '.join(missing)}")
        return False
    return True


def convert_taxonomy_rows(rows: list[dict]) -> tuple[list[dict], int, int]:
    """
    Convert all_valid_cases_with_taxonomy.csv rows to the internal schema.

    Simple-heuristics confidence is similarity_score.

    Returns:
        converted rows,
        rows filtered because taxonomy_label is different_unclassified,
        rows filtered because taxonomy_label is Same.
    """
    converted = []
    unclassified_filtered_out = 0
    same_filtered_out = 0

    for row in rows:
        raw_label = row.get("taxonomy_label", "")

        if not should_keep_simple_heuristics_row(raw_label):
            unclassified_filtered_out += 1
            continue

        normalized_label = normalize_taxonomy_label(raw_label)
        if not is_inconsistent_taxonomy_label(normalized_label):
            same_filtered_out += 1
            continue

        file_path = normalize_file_path_key(row.get("file_path", ""))
        similarity_score = row.get("similarity_score", "")
        similarity_score_float = parse_float(similarity_score)

        converted.append(
            {
                "question": row.get("question", ""),
                "gold_answer": row.get("gold_answer", ""),
                "KG answer": row.get("result_cleaned", ""),
                "taxonomy_label": normalized_label,
                "sparql": row.get("sparql", ""),
                "confidence": similarity_score,
                "method": SIMPLE_METHOD,
                "source": file_path,
                "_file_path": file_path,
                "_qa_group": qa_group_from_source(file_path),
                "_similarity_score": similarity_score_float,
                "_difference_severity": "",
            }
        )

    return converted, unclassified_filtered_out, same_filtered_out


def build_sparql_lookup(taxonomy_rows: list[dict]) -> tuple[dict[str, str], dict[str, str]]:
    """
    Build SPARQL lookup dictionaries from all_valid_cases_with_taxonomy.csv.

    The file_path lookup is preferred because it is the most stable key across
    step 1 inputs. The question lookup is a fallback for rows where file_path is
    missing or formatted differently.
    """
    sparql_by_file_path = {}
    sparql_by_question = {}

    for row in taxonomy_rows:
        sparql = row.get("sparql", "")

        file_path = normalize_file_path_key(row.get("file_path", ""))
        if file_path and file_path not in sparql_by_file_path:
            sparql_by_file_path[file_path] = sparql

        question = (row.get("question", "") or "").strip()
        if question and question not in sparql_by_question:
            sparql_by_question[question] = sparql

    return sparql_by_file_path, sparql_by_question


def lookup_sparql_for_row(
    row: dict,
    sparql_by_file_path: dict[str, str],
    sparql_by_question: dict[str, str],
) -> str:
    """Return the SPARQL query for a row using direct value, file_path, then question."""
    direct_sparql = row.get("sparql", "")
    if direct_sparql:
        return direct_sparql

    file_path = normalize_file_path_key(row.get("file_path", ""))
    if file_path in sparql_by_file_path:
        return sparql_by_file_path[file_path]

    question = (row.get("question", "") or "").strip()
    if question in sparql_by_question:
        return sparql_by_question[question]

    return ""


def convert_judged_rows(
    rows: list[dict],
    sparql_by_file_path: dict[str, str] | None = None,
    sparql_by_question: dict[str, str] | None = None,
) -> tuple[list[dict], int]:
    """
    Convert *_judged_filtered.csv rows to the internal schema.

    LLM-as-a-judge confidence is difference_severity. SPARQL is copied from the
    judged row when present, otherwise looked up from all_valid_cases_with_taxonomy.csv.

    Returns:
        converted rows,
        rows filtered because taxonomy_label is Same.
    """
    converted = []
    same_filtered_out = 0
    sparql_by_file_path = sparql_by_file_path or {}
    sparql_by_question = sparql_by_question or {}

    for row in rows:
        normalized_label = normalize_taxonomy_label(row.get("taxonomy_label", ""))
        if not is_inconsistent_taxonomy_label(normalized_label):
            same_filtered_out += 1
            continue

        file_path = normalize_file_path_key(row.get("file_path", ""))
        difference_severity = normalize_difference_severity(row.get("difference_severity", ""))

        converted.append(
            {
                "question": row.get("question", ""),
                "gold_answer": row.get("gold_answer", ""),
                "KG answer": row.get("KG answer", ""),
                "taxonomy_label": normalized_label,
                "sparql": lookup_sparql_for_row(
                    row,
                    sparql_by_file_path,
                    sparql_by_question,
                ),
                "confidence": difference_severity,
                "method": LLM_METHOD,
                "source": file_path,
                "_file_path": file_path,
                "_qa_group": qa_group_from_source(file_path),
                "_similarity_score": math.nan,
                "_difference_severity": difference_severity,
            }
        )

    return converted, same_filtered_out


def public_row(row: dict) -> dict:
    """Return only columns intended for public output CSVs."""
    return {column: row.get(column, "") for column in OUTPUT_COLUMNS}


def filter_large_result_rows(rows: list[dict], large_file_paths: set[str]) -> tuple[list[dict], int]:
    """Remove rows whose source/file_path appears in the >10-row report."""
    kept_rows = []
    removed_count = 0

    for row in rows:
        if row.get("_file_path", "") in large_file_paths:
            removed_count += 1
        else:
            kept_rows.append(row)

    return kept_rows, removed_count


def deduplicate_by_question(rows: list[dict]) -> tuple[list[dict], int]:
    """Remove duplicate questions, keeping the first occurrence."""
    seen_questions = set()
    kept_rows = []
    removed_count = 0

    for row in rows:
        question_key = (row.get("question", "") or "").strip()
        if question_key in seen_questions:
            removed_count += 1
        else:
            seen_questions.add(question_key)
            kept_rows.append(row)

    return kept_rows, removed_count


def row_confidence_sort_key(row: dict) -> tuple:
    """
    Return a method-specific confidence sort key.

    Simple-heuristics rows are sorted by similarity_score, high score first.
    LLM-as-a-judge rows are sorted by difference_severity:
    major, moderate, minor, none.
    """
    method = row.get("method", "")

    if method == SIMPLE_METHOD:
        return (
            -safe_sort_float(row.get("_similarity_score", math.nan)),
            row.get("question", ""),
        )

    if method == LLM_METHOD:
        severity = row.get("_difference_severity", "")
        severity_rank = DIFFERENCE_SEVERITY_ORDER.get(severity, len(DIFFERENCE_SEVERITY_ORDER))
        return (
            severity_rank,
            row.get("question", ""),
        )

    return (
        999,
        row.get("question", ""),
    )


def bucket_rows(rows: list[dict]) -> dict[tuple[str, str], list[dict]]:
    """Bucket rows by QA group and method, then sort each bucket by confidence."""
    buckets: dict[tuple[str, str], list[dict]] = defaultdict(list)

    for row in rows:
        qa_group = row.get("_qa_group", "")
        method = row.get("method", "")
        buckets[(qa_group, method)].append(row)

    for key in list(buckets):
        buckets[key].sort(key=row_confidence_sort_key)

    return buckets


def bucket_count_summary(buckets: dict[tuple[str, str], list[dict]]) -> dict[tuple[str, str], int]:
    """Return candidate counts for each source/method bucket."""
    return {
        (SIMPLE_QA_GROUP, SIMPLE_METHOD): len(buckets.get((SIMPLE_QA_GROUP, SIMPLE_METHOD), [])),
        (SIMPLE_QA_GROUP, LLM_METHOD): len(buckets.get((SIMPLE_QA_GROUP, LLM_METHOD), [])),
        (COMPLEX_QA_GROUP, SIMPLE_METHOD): len(buckets.get((COMPLEX_QA_GROUP, SIMPLE_METHOD), [])),
        (COMPLEX_QA_GROUP, LLM_METHOD): len(buckets.get((COMPLEX_QA_GROUP, LLM_METHOD), [])),
    }


def maximum_perfect_balanced_size(buckets: dict[tuple[str, str], list[dict]], requested_size: int) -> int:
    """
    Return the maximum perfect 2x2-balanced size up to the requested size.

    Perfect 2x2 balance means all four source/method buckets have equal count.
    Therefore the final size must be a multiple of 4.
    """
    counts = bucket_count_summary(buckets)
    smallest_bucket = min(counts.values())
    requested_per_bucket = requested_size // 4

    possible_per_bucket = min(smallest_bucket, requested_per_bucket)
    return possible_per_bucket * 4


def is_exact_perfect_balance_possible(
    buckets: dict[tuple[str, str], list[dict]],
    requested_size: int,
) -> bool:
    """Return True if the requested target can be exactly 2x2 balanced."""
    if requested_size % 4 != 0:
        return False

    counts = bucket_count_summary(buckets)
    required_per_bucket = requested_size // 4
    return all(count >= required_per_bucket for count in counts.values())


def prompt_for_perfect_balance(requested_size: int, max_perfect_size: int) -> bool:
    """
    Ask whether to use the maximum perfect-balanced set.

    Returns True if the user enters y, otherwise False.
    """
    print()
    print("=" * 72)
    print("Perfect 2x2 balance check")
    print("=" * 72)
    print(f"Requested main target size: {requested_size}")
    print(f"Maximum possible perfect-balanced size: {max_perfect_size}")
    print()
    print(
        "The requested target cannot be perfectly balanced across "
        "SimpleQA/ComplexQA and Simple heuristics/LLM-as-a-judge."
    )
    print()
    print("Enter y to generate the maximum perfect-balanced set.")
    print("Enter n to generate the requested target size using source-first balancing.")
    print()

    while True:
        answer = input("Do you want to generate the maximum perfect-balanced set? [y/n]: ").strip().lower()
        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False
        print("Please enter y or n.")


def build_exact_perfect_quotas(selection_size: int) -> dict[tuple[str, str], int]:
    """Build exact 2x2 quotas for a perfect-balanced selection."""
    if selection_size % 4 != 0:
        raise ValueError("Perfect-balanced selection size must be divisible by 4.")

    per_bucket = selection_size // 4
    return {
        (SIMPLE_QA_GROUP, SIMPLE_METHOD): per_bucket,
        (SIMPLE_QA_GROUP, LLM_METHOD): per_bucket,
        (COMPLEX_QA_GROUP, SIMPLE_METHOD): per_bucket,
        (COMPLEX_QA_GROUP, LLM_METHOD): per_bucket,
    }


def build_source_quotas(selection_size: int) -> dict[str, int]:
    """Build exact SimpleQA/ComplexQA quotas."""
    if selection_size % 2 != 0:
        raise ValueError("Source-balanced selection size must be even.")

    per_source = selection_size // 2
    return {
        SIMPLE_QA_GROUP: per_source,
        COMPLEX_QA_GROUP: per_source,
    }


def build_method_preferences_for_source_first(selection_size: int) -> dict[str, dict[str, int]]:
    """
    Build preferred method quotas inside each source group.

    The quotas are preferences. If one method bucket is short, the deficit is
    complemented from the other method inside the same source group.
    """
    source_quotas = build_source_quotas(selection_size)

    preferences = {}
    for qa_group, source_quota in source_quotas.items():
        simple_quota = source_quota // 2
        llm_quota = source_quota - simple_quota
        preferences[qa_group] = {
            SIMPLE_METHOD: simple_quota,
            LLM_METHOD: llm_quota,
        }

    return preferences


def second10_source_quotas() -> dict[str, int]:
    """Build source quotas for the second 10 presentation slice."""
    return {
        SIMPLE_QA_GROUP: 5,
        COMPLEX_QA_GROUP: 5,
    }


def second10_method_preferences() -> dict[str, dict[str, int]]:
    """
    Build preferred method quotas for the second 10 presentation slice.

    Source balance is exact. Method balance is preferred but can be complemented.
    """
    return {
        SIMPLE_QA_GROUP: {
            SIMPLE_METHOD: 3,
            LLM_METHOD: 2,
        },
        COMPLEX_QA_GROUP: {
            SIMPLE_METHOD: 2,
            LLM_METHOD: 3,
        },
    }


def take_from_bucket(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    key: tuple[str, str],
    count: int,
) -> list[dict]:
    """Take the next count rows from one sorted bucket."""
    if count <= 0:
        return []

    bucket = buckets.get(key, [])
    start = used_counts.get(key, 0)
    end = min(start + count, len(bucket))

    selected = bucket[start:end]
    used_counts[key] = end

    return selected


def remaining_in_bucket(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    key: tuple[str, str],
) -> int:
    """Return the number of unused rows left in one bucket."""
    return max(0, len(buckets.get(key, [])) - used_counts.get(key, 0))


def validate_exact_bucket_capacity(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    quotas: dict[tuple[str, str], int],
) -> None:
    """Check that every exact source/method bucket quota can be satisfied."""
    shortages = []

    for key, quota in quotas.items():
        available = remaining_in_bucket(buckets, used_counts, key)
        if available < quota:
            qa_group, method = key
            shortages.append(
                f"{qa_group} / {method}: required {quota}, available {available}"
            )

    if shortages:
        raise RuntimeError(
            "Not enough rows for exact perfect-balanced selection:\n  - "
            + "\n  - ".join(shortages)
        )


def validate_source_capacity(
    buckets: dict[tuple[str, str], list[dict]],
    source_quotas: dict[str, int],
    used_counts: dict[tuple[str, str], int],
) -> None:
    """
    Check that each source group has enough remaining rows.

    This enforces the main fallback priority: SimpleQA and ComplexQA must stay
    balanced exactly. Method balance is preferred but not enforced if one method
    bucket is short.
    """
    shortages = []

    for qa_group, source_quota in source_quotas.items():
        available = 0
        for method in [SIMPLE_METHOD, LLM_METHOD]:
            key = (qa_group, method)
            available += remaining_in_bucket(buckets, used_counts, key)

        if available < source_quota:
            shortages.append(
                f"{qa_group}: required {source_quota}, available {available}"
            )

    if shortages:
        raise RuntimeError(
            "Not enough eligible inconsistent rows to keep exact SimpleQA / "
            "ComplexQA balance:\n  - " + "\n  - ".join(shortages)
        )


def select_exact_perfect_slice(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    quotas: dict[tuple[str, str], int],
) -> list[dict]:
    """Select one exact 2x2-balanced slice."""
    validate_exact_bucket_capacity(
        buckets=buckets,
        used_counts=used_counts,
        quotas=quotas,
    )

    selected = []
    for key in [
        (SIMPLE_QA_GROUP, SIMPLE_METHOD),
        (SIMPLE_QA_GROUP, LLM_METHOD),
        (COMPLEX_QA_GROUP, SIMPLE_METHOD),
        (COMPLEX_QA_GROUP, LLM_METHOD),
    ]:
        selected.extend(take_from_bucket(buckets, used_counts, key, quotas[key]))

    return interleave_balanced_rows(selected)


def select_source_group_rows(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    qa_group: str,
    source_quota: int,
    method_preferences: dict[str, int],
) -> tuple[list[dict], list[str]]:
    """
    Select rows for one source group.

    This function prioritizes the source quota. It first tries to follow the
    preferred method split. If one method is short, it fills the deficit from
    the other method inside the same source group.
    """
    selected = []
    notes = []

    simple_key = (qa_group, SIMPLE_METHOD)
    llm_key = (qa_group, LLM_METHOD)

    simple_preference = method_preferences.get(SIMPLE_METHOD, 0)
    llm_preference = method_preferences.get(LLM_METHOD, 0)

    simple_available = remaining_in_bucket(buckets, used_counts, simple_key)
    llm_available = remaining_in_bucket(buckets, used_counts, llm_key)

    simple_take = min(simple_preference, simple_available)
    llm_take = min(llm_preference, llm_available)

    selected.extend(take_from_bucket(buckets, used_counts, simple_key, simple_take))
    selected.extend(take_from_bucket(buckets, used_counts, llm_key, llm_take))

    deficit = source_quota - len(selected)

    if deficit > 0:
        simple_remaining = remaining_in_bucket(buckets, used_counts, simple_key)
        llm_remaining = remaining_in_bucket(buckets, used_counts, llm_key)

        if simple_take < simple_preference:
            fill_method = LLM_METHOD
            fill_key = llm_key
            fill_available = llm_remaining
        elif llm_take < llm_preference:
            fill_method = SIMPLE_METHOD
            fill_key = simple_key
            fill_available = simple_remaining
        else:
            if simple_remaining >= llm_remaining:
                fill_method = SIMPLE_METHOD
                fill_key = simple_key
                fill_available = simple_remaining
            else:
                fill_method = LLM_METHOD
                fill_key = llm_key
                fill_available = llm_remaining

        fill_count = min(deficit, fill_available)
        selected.extend(take_from_bucket(buckets, used_counts, fill_key, fill_count))
        deficit -= fill_count

        if fill_count > 0:
            notes.append(
                f"{qa_group}: filled {fill_count} row(s) from {fill_method} "
                "because the preferred method bucket was short."
            )

    if deficit > 0:
        for method in [SIMPLE_METHOD, LLM_METHOD]:
            key = (qa_group, method)
            fill_count = min(deficit, remaining_in_bucket(buckets, used_counts, key))
            selected.extend(take_from_bucket(buckets, used_counts, key, fill_count))
            deficit -= fill_count
            if deficit == 0:
                break

    if len(selected) != source_quota:
        raise RuntimeError(
            f"Could not select {source_quota} rows for {qa_group}; "
            f"selected only {len(selected)}."
        )

    return selected, notes


def select_source_first_slice(
    buckets: dict[tuple[str, str], list[dict]],
    used_counts: dict[tuple[str, str], int],
    source_quotas: dict[str, int],
    method_preferences: dict[str, dict[str, int]],
) -> tuple[list[dict], list[str]]:
    """
    Select one source-balanced slice.

    SimpleQA and ComplexQA quotas are exact. Method quotas are preferred, but
    shortages are complemented within the same source group.
    """
    validate_source_capacity(
        buckets=buckets,
        source_quotas=source_quotas,
        used_counts=used_counts,
    )

    selected = []
    notes = []

    for qa_group in [SIMPLE_QA_GROUP, COMPLEX_QA_GROUP]:
        group_rows, group_notes = select_source_group_rows(
            buckets=buckets,
            used_counts=used_counts,
            qa_group=qa_group,
            source_quota=source_quotas[qa_group],
            method_preferences=method_preferences[qa_group],
        )
        selected.extend(group_rows)
        notes.extend(group_notes)

    selected = interleave_balanced_rows(selected)
    return selected, notes


def choose_main_selection_mode(
    buckets: dict[tuple[str, str], list[dict]],
    requested_size: int,
) -> tuple[str, int]:
    """
    Decide whether to use exact perfect balance or source-first fallback.

    Returns:
        mode, main_selection_size

    mode is one of:
        perfect
        source_first
    """
    print_bucket_scan(buckets)

    exact_possible = is_exact_perfect_balance_possible(
        buckets=buckets,
        requested_size=requested_size,
    )
    max_perfect_size = maximum_perfect_balanced_size(
        buckets=buckets,
        requested_size=requested_size,
    )

    if exact_possible:
        print()
        print("=" * 72)
        print("Perfect 2x2 balance check")
        print("=" * 72)
        print(f"Requested main target size: {requested_size}")
        print("Exact perfect balance is possible.")
        return "perfect", requested_size

    use_perfect = prompt_for_perfect_balance(
        requested_size=requested_size,
        max_perfect_size=max_perfect_size,
    )

    if use_perfect:
        return "perfect", max_perfect_size

    return "source_first", requested_size


def print_bucket_scan(buckets: dict[tuple[str, str], list[dict]]) -> None:
    """Print candidate counts before deciding the balancing mode."""
    counts = bucket_count_summary(buckets)

    print("=" * 72)
    print("Eligible inconsistent candidate scan")
    print("=" * 72)
    print(f"{SIMPLE_QA_GROUP} / {SIMPLE_METHOD}: {counts[(SIMPLE_QA_GROUP, SIMPLE_METHOD)]}")
    print(f"{SIMPLE_QA_GROUP} / {LLM_METHOD}: {counts[(SIMPLE_QA_GROUP, LLM_METHOD)]}")
    print(f"{COMPLEX_QA_GROUP} / {SIMPLE_METHOD}: {counts[(COMPLEX_QA_GROUP, SIMPLE_METHOD)]}")
    print(f"{COMPLEX_QA_GROUP} / {LLM_METHOD}: {counts[(COMPLEX_QA_GROUP, LLM_METHOD)]}")
    print()


def select_main_and_second_slice(
    rows: list[dict],
    requested_size: int,
) -> tuple[list[dict], list[dict], str, int, list[str]]:
    """
    Select the main annotation set and then the next 10 presentation cases.

    The main set is either:
    - exact perfect 2x2-balanced, or
    - source-first fallback, depending on the interactive decision.

    The second 10 is always source-balanced. Method balance is preferred, but
    method shortages are complemented within the same source group.
    """
    buckets = bucket_rows(rows)
    mode, main_selection_size = choose_main_selection_mode(
        buckets=buckets,
        requested_size=requested_size,
    )

    used_counts: dict[tuple[str, str], int] = defaultdict(int)
    notes = []

    if mode == "perfect":
        main_rows = select_exact_perfect_slice(
            buckets=buckets,
            used_counts=used_counts,
            quotas=build_exact_perfect_quotas(main_selection_size),
        )
    else:
        main_rows, main_notes = select_source_first_slice(
            buckets=buckets,
            used_counts=used_counts,
            source_quotas=build_source_quotas(main_selection_size),
            method_preferences=build_method_preferences_for_source_first(main_selection_size),
        )
        notes.extend(main_notes)

    second_rows, second_notes = select_source_first_slice(
        buckets=buckets,
        used_counts=used_counts,
        source_quotas=second10_source_quotas(),
        method_preferences=second10_method_preferences(),
    )
    notes.extend(second_notes)

    return main_rows, second_rows, mode, main_selection_size, notes


def interleave_balanced_rows(rows: list[dict]) -> list[dict]:
    """
    Interleave rows from QA/method buckets.

    This keeps the final CSV balanced throughout the file instead of writing one
    large method or QA block. Rows inside each bucket already follow the correct
    confidence order.
    """
    buckets = bucket_rows(rows)

    bucket_order = [
        (SIMPLE_QA_GROUP, SIMPLE_METHOD),
        (COMPLEX_QA_GROUP, LLM_METHOD),
        (SIMPLE_QA_GROUP, LLM_METHOD),
        (COMPLEX_QA_GROUP, SIMPLE_METHOD),
    ]

    output = []
    bucket_positions = {key: 0 for key in bucket_order}

    while len(output) < len(rows):
        progressed = False

        for key in bucket_order:
            position = bucket_positions[key]
            bucket = buckets.get(key, [])
            if position < len(bucket):
                output.append(bucket[position])
                bucket_positions[key] += 1
                progressed = True

        if not progressed:
            break

    return output


def write_csv(csv_path: Path, rows: list[dict]) -> None:
    """Write final annotation rows to a CSV file."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(public_row(row) for row in rows)


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    """Build a Markdown table."""
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(lines)


def count_by_label(rows: list[dict]) -> Counter:
    """Count normalized taxonomy labels."""
    return Counter(normalize_taxonomy_label(row.get("taxonomy_label", "")) or "UNKNOWN" for row in rows)


def count_by_method(rows: list[dict]) -> Counter:
    """Count selected rows by internal method."""
    return Counter(row.get("method", "") or "UNKNOWN" for row in rows)


def count_by_qa_group(rows: list[dict]) -> Counter:
    """Count selected rows by QA group."""
    return Counter(row.get("_qa_group", "") or "UNKNOWN" for row in rows)


def count_by_bucket(rows: list[dict]) -> Counter:
    """Count selected rows by QA group and method."""
    return Counter((row.get("_qa_group", "UNKNOWN"), row.get("method", "UNKNOWN")) for row in rows)


def build_distribution_tables(rows: list[dict]) -> list[str]:
    """Build Markdown distribution tables for one selected output."""
    lines = []

    method_counts = count_by_method(rows)
    qa_group_counts = count_by_qa_group(rows)
    bucket_counts = count_by_bucket(rows)
    taxonomy_counts = count_by_label(rows)

    lines.append("Method distribution:")
    lines.append("")
    lines.append(
        markdown_table(
            ["Method", "Count"],
            [
                [SIMPLE_METHOD, str(method_counts.get(SIMPLE_METHOD, 0))],
                [LLM_METHOD, str(method_counts.get(LLM_METHOD, 0))],
            ],
        )
    )
    lines.append("")

    lines.append("QA group distribution:")
    lines.append("")
    lines.append(
        markdown_table(
            ["QA Group", "Count"],
            [
                [SIMPLE_QA_GROUP, str(qa_group_counts.get(SIMPLE_QA_GROUP, 0))],
                [COMPLEX_QA_GROUP, str(qa_group_counts.get(COMPLEX_QA_GROUP, 0))],
            ],
        )
    )
    lines.append("")

    lines.append("2x2 bucket distribution:")
    lines.append("")
    lines.append(
        markdown_table(
            ["QA Group", "Method", "Count"],
            [
                [SIMPLE_QA_GROUP, SIMPLE_METHOD, str(bucket_counts.get((SIMPLE_QA_GROUP, SIMPLE_METHOD), 0))],
                [SIMPLE_QA_GROUP, LLM_METHOD, str(bucket_counts.get((SIMPLE_QA_GROUP, LLM_METHOD), 0))],
                [COMPLEX_QA_GROUP, SIMPLE_METHOD, str(bucket_counts.get((COMPLEX_QA_GROUP, SIMPLE_METHOD), 0))],
                [COMPLEX_QA_GROUP, LLM_METHOD, str(bucket_counts.get((COMPLEX_QA_GROUP, LLM_METHOD), 0))],
            ],
        )
    )
    lines.append("")

    lines.append("Taxonomy distribution:")
    lines.append("")
    lines.append(
        markdown_table(
            ["taxonomy_label", "Count"],
            [[label, str(count)] for label, count in sorted(taxonomy_counts.items())],
        )
    )
    lines.append("")

    return lines


def write_statistics(
    statistics_path: Path,
    main_rows: list[dict],
    second_rows: list[dict],
    cleaning_summary: dict,
    selection_mode: str,
    main_selection_size: int,
    selection_notes: list[str],
) -> None:
    """Write Markdown statistics for the 235B annotation outputs."""
    statistics_path.parent.mkdir(parents=True, exist_ok=True)

    lines = []
    lines.append("# 235B Interactive Inconsistent Annotation Selection Statistics")
    lines.append("")
    lines.append("This file summarizes the balanced 235B inconsistent-case annotation outputs.")
    lines.append("")
    lines.append("Rows whose normalized `taxonomy_label` is `Same` are excluded before selection.")
    lines.append("")
    lines.append("Public output columns:")
    lines.append("")
    lines.append("```text")
    lines.append(",".join(OUTPUT_COLUMNS))
    lines.append("```")
    lines.append("")

    lines.append("## Selection Mode")
    lines.append("")
    lines.append(f"- Requested target size: {TARGET_SELECTION_CASES}")
    lines.append(f"- Actual main selection size: {main_selection_size}")
    lines.append(f"- Mode: {selection_mode}")
    lines.append("")
    if selection_mode == "perfect":
        lines.append("The main set is exactly balanced across the four source/method buckets.")
    else:
        lines.append("The main set uses source-first balancing with method complementing.")
    lines.append("")

    lines.append("## Cleaning Summary")
    lines.append("")
    lines.append(
        markdown_table(
            ["Metric", "Count"],
            [
                ["Simple heuristics input rows", str(cleaning_summary["taxonomy_input_rows"])],
                ["LLM-as-a-judge input rows", str(cleaning_summary["judged_input_rows"])],
                ["Simple rows filtered by different_unclassified", str(cleaning_summary["filtered_unclassified_simple_rows"])],
                ["Simple rows filtered because taxonomy_label is Same", str(cleaning_summary["filtered_same_simple_rows"])],
                ["LLM rows filtered because taxonomy_label is Same", str(cleaning_summary["filtered_same_judged_rows"])],
                ["Rows removed by >10-row file_path filter", str(cleaning_summary["large_result_rows_removed"])],
                ["Duplicate question rows removed", str(cleaning_summary["duplicate_question_rows_removed"])],
                ["Eligible inconsistent rows after cleaning", str(cleaning_summary["eligible_rows"])],
            ],
        )
    )
    lines.append("")

    lines.append("## Selection Notes")
    lines.append("")
    if selection_notes:
        for note in selection_notes:
            lines.append(f"- {note}")
    else:
        lines.append("No method-bucket shortages were encountered.")
    lines.append("")

    lines.append("## Main Inconsistent Output")
    lines.append("")
    lines.append(f"Rows written: {len(main_rows)}")
    lines.append("")
    lines.extend(build_distribution_tables(main_rows))

    lines.append("## Second 10-Case Inconsistent Presentation Slice")
    lines.append("")
    lines.append(f"Rows written: {len(second_rows)}")
    lines.append("")
    lines.extend(build_distribution_tables(second_rows))

    with statistics_path.open("w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def print_counts(label: str, rows: list[dict]) -> None:
    """Print method, QA group, 2x2 bucket, and taxonomy counts."""
    method_counts = count_by_method(rows)
    qa_group_counts = count_by_qa_group(rows)
    bucket_counts = count_by_bucket(rows)
    taxonomy_counts = count_by_label(rows)

    print(f"  {label}:")
    print(f"    Rows: {len(rows)}")
    print("    Method counts:")
    print(f"      {SIMPLE_METHOD}: {method_counts.get(SIMPLE_METHOD, 0)}")
    print(f"      {LLM_METHOD}: {method_counts.get(LLM_METHOD, 0)}")
    print("    QA group counts:")
    print(f"      {SIMPLE_QA_GROUP}: {qa_group_counts.get(SIMPLE_QA_GROUP, 0)}")
    print(f"      {COMPLEX_QA_GROUP}: {qa_group_counts.get(COMPLEX_QA_GROUP, 0)}")
    print("    2x2 bucket counts:")
    print(f"      {SIMPLE_QA_GROUP} / {SIMPLE_METHOD}: {bucket_counts.get((SIMPLE_QA_GROUP, SIMPLE_METHOD), 0)}")
    print(f"      {SIMPLE_QA_GROUP} / {LLM_METHOD}: {bucket_counts.get((SIMPLE_QA_GROUP, LLM_METHOD), 0)}")
    print(f"      {COMPLEX_QA_GROUP} / {SIMPLE_METHOD}: {bucket_counts.get((COMPLEX_QA_GROUP, SIMPLE_METHOD), 0)}")
    print(f"      {COMPLEX_QA_GROUP} / {LLM_METHOD}: {bucket_counts.get((COMPLEX_QA_GROUP, LLM_METHOD), 0)}")
    print("    Taxonomy counts:")
    for taxonomy_label, count in sorted(taxonomy_counts.items()):
        print(f"      {taxonomy_label}: {count}")


def run_step1_build_annotation_csvs() -> None:
    """Run step 1: build the 235B-only inconsistent annotation CSV files."""
    try:
        project_root = find_project_root()
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return

    print(f"Project root: {project_root}")
    print()

    pair = MODEL_PAIR
    model = pair["model"]
    judged_csv = project_root / pair["judged_csv"]
    taxonomy_csv = project_root / pair["taxonomy_csv"]
    main_csv = project_root / pair["main_csv"]
    second10_csv = project_root / pair["second10_csv"]
    statistics_md = project_root / pair["statistics_md"]

    missing_files = []
    if not judged_csv.exists():
        missing_files.append(str(judged_csv))
    if not taxonomy_csv.exists():
        missing_files.append(str(taxonomy_csv))

    if missing_files:
        print(f"ERROR: Missing file(s) for {model}:")
        for missing in missing_files:
            print(f"  - {missing}")
        return

    large_results_report = project_root / LARGE_RESULTS_REPORT
    large_paths_by_model = parse_large_results_report(large_results_report)

    model_output_dir = pair["taxonomy_csv"].parent.name
    large_file_paths = large_paths_by_model.get(model_output_dir, set())

    print(f"Processing {model}")
    print(f"  Taxonomy CSV: {taxonomy_csv}")
    print(f"  Judged CSV:   {judged_csv}")
    print()

    taxonomy_rows = read_csv_rows(taxonomy_csv)
    judged_rows = read_csv_rows(judged_csv)

    taxonomy_required = [
        "question",
        "gold_answer",
        "result_cleaned",
        "taxonomy_label",
        "sparql",
        "similarity_score",
        "file_path",
    ]
    judged_required = [
        "question",
        "gold_answer",
        "KG answer",
        "taxonomy_label",
        "difference_severity",
        "file_path",
    ]

    taxonomy_ok = require_columns(taxonomy_csv, taxonomy_rows, taxonomy_required)
    judged_ok = require_columns(judged_csv, judged_rows, judged_required)
    if not taxonomy_ok or not judged_ok:
        print("ERROR: Required columns are missing.")
        return

    try:
        (
            converted_taxonomy_rows,
            unclassified_simple_filtered_count,
            same_simple_filtered_count,
        ) = convert_taxonomy_rows(taxonomy_rows)
        sparql_by_file_path, sparql_by_question = build_sparql_lookup(taxonomy_rows)
        converted_judged_rows, same_judged_filtered_count = convert_judged_rows(
            judged_rows,
            sparql_by_file_path=sparql_by_file_path,
            sparql_by_question=sparql_by_question,
        )
    except ValueError as exc:
        print(f"ERROR: {exc}")
        return

    taxonomy_after_large_filter, taxonomy_large_removed = filter_large_result_rows(
        converted_taxonomy_rows,
        large_file_paths,
    )
    judged_after_large_filter, judged_large_removed = filter_large_result_rows(
        converted_judged_rows,
        large_file_paths,
    )

    long_answer_removed_count = taxonomy_large_removed + judged_large_removed

    merged_rows_before_dedup = taxonomy_after_large_filter + judged_after_large_filter
    merged_rows, duplicate_removed_count = deduplicate_by_question(merged_rows_before_dedup)

    try:
        (
            main_rows,
            second10_rows,
            selection_mode,
            main_selection_size,
            selection_notes,
        ) = select_main_and_second_slice(
            rows=merged_rows,
            requested_size=TARGET_SELECTION_CASES,
        )
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return

    write_csv(main_csv, main_rows)
    write_csv(second10_csv, second10_rows)

    cleaning_summary = {
        "taxonomy_input_rows": len(taxonomy_rows),
        "judged_input_rows": len(judged_rows),
        "filtered_unclassified_simple_rows": unclassified_simple_filtered_count,
        "filtered_same_simple_rows": same_simple_filtered_count,
        "filtered_same_judged_rows": same_judged_filtered_count,
        "large_result_rows_removed": long_answer_removed_count,
        "duplicate_question_rows_removed": duplicate_removed_count,
        "eligible_rows": len(merged_rows),
    }

    write_statistics(
        statistics_path=statistics_md,
        main_rows=main_rows,
        second_rows=second10_rows,
        cleaning_summary=cleaning_summary,
        selection_mode=selection_mode,
        main_selection_size=main_selection_size,
        selection_notes=selection_notes,
    )

    print()
    print("=" * 72)
    print("Cleaning summary")
    print("=" * 72)
    print(f"  Simple heuristics input rows: {len(taxonomy_rows)}")
    print(f"  LLM-as-a-judge input rows: {len(judged_rows)}")
    print(f"  Simple rows filtered by different_unclassified: {unclassified_simple_filtered_count}")
    print(f"  Simple rows filtered because taxonomy_label is Same: {same_simple_filtered_count}")
    print(f"  LLM rows filtered because taxonomy_label is Same: {same_judged_filtered_count}")
    print(f"  Rows removed by >10-row file_path filter: {long_answer_removed_count}")
    print(f"    - Simple heuristics removed: {taxonomy_large_removed}")
    print(f"    - LLM-as-a-judge removed:   {judged_large_removed}")
    print(f"  Duplicate question rows removed: {duplicate_removed_count}")
    print(f"  Eligible inconsistent rows after cleaning: {len(merged_rows)}")
    print()

    print("=" * 72)
    print("Selection summary")
    print("=" * 72)
    print(f"  Requested main target size: {TARGET_SELECTION_CASES}")
    print(f"  Actual main selection size: {main_selection_size}")
    print(f"  Selection mode: {selection_mode}")
    print()

    if selection_notes:
        print("Selection notes:")
        for note in selection_notes:
            print(f"  - {note}")
        print()

    print_counts("Main inconsistent annotation cases", main_rows)
    print()
    print_counts("Second 10 balanced inconsistent presentation cases", second10_rows)
    print()

    print("Outputs written:")
    print(f"  Main CSV:       {main_csv}")
    print(f"  Second 10 CSV:  {second10_csv}")
    print(f"  Statistics MD:  {statistics_md}")
    print()
    print("Done.")


# =============================================================================
# Step 2: split the generated CSV into formatted Excel annotation groups.
# =============================================================================

# pip install pandas openpyxl

"""
split.py

Purpose
-------
Convert an input CSV annotation file into multiple balanced Excel annotation files.

Input CSV columns expected:
    question,gold_answer,KG answer,taxonomy_label,sparql,confidence,source

Output:
    annotation_outputs/annotation_group_1.xlsx
    annotation_outputs/annotation_group_2.xlsx
    annotation_outputs/annotation_group_3.xlsx

Each output file:
    - removes the confidence column
    - keeps the SPARQL column for review
    - adds an empty Corrected SPARQL column for cases where SPARQL needs correction
    - keeps a stable case_id for merging later
    - adds dropdown column: label_correctness
    - adds three Yes/No columns for Wikidata structural causes:
        missing_edge
        missing_node
        missing_property_or_qualifier
    - adds free-text column: note
    - uses alternating row colors for better readability

Important
---------
CSV files cannot store dropdown menus.
Therefore, this script outputs .xlsx files instead of .csv files.
"""

import pandas as pd
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Dropdown options
# ---------------------------------------------------------------------

# Annotators use this column to verify whether the assigned taxonomy label
# is correct according to answer-level comparison.
TAXONOMY_LABEL_CORRECT_OPTIONS = [
    "Correct",
    "Incorrect",
    "Unsure"
]

# Annotators use these values for each Wikidata structural cause column.
# Each cause is now annotated independently, so multiple causes can be marked.
YES_NO_OPTIONS = [
    "Yes",
    "No"
]


# ---------------------------------------------------------------------
# Helper 1: clean column names
# ---------------------------------------------------------------------

def clean_column_names(df):
    """
    Clean CSV column names.

    Why this is needed:
    Sometimes CSV files contain invisible characters such as BOM '\\ufeff',
    or spaces around column names.

    Example:
        ' taxonomy_label ' -> 'taxonomy_label'
        '\\ufeffquestion' -> 'question'
    """

    df.columns = (
        df.columns
        .astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
    )

    return df


# ---------------------------------------------------------------------
# Helper 2: balanced splitting
# ---------------------------------------------------------------------

def split_balanced(df, n_groups=3, stratify_cols=None):
    """
    Split a dataframe into n balanced groups.

    Parameters
    ----------
    df : pandas.DataFrame
        Full dataframe to split.

    n_groups : int
        Number of output groups.

    stratify_cols : list[str] or None
        Columns used for approximate stratified splitting.

        Recommended:
            stratify_cols=["taxonomy_label"]

        Not recommended:
            stratify_cols=["source"]

        Why not use source?
            In your file, source may be unique or almost unique for every row.
            If we group by source directly, each source group has only one row.
            A bad splitting function may then put every row into group 1.

    Returns
    -------
    list[pandas.DataFrame]
        A list of n_groups dataframes.
    """

    df = df.reset_index(drop=True).copy()

    groups = [[] for _ in range(n_groups)]

    if stratify_cols:
        for col in stratify_cols:
            if col not in df.columns:
                raise ValueError(
                    f"Stratification column not found: {col}\n"
                    f"Available columns: {list(df.columns)}"
                )

        grouped = df.groupby(stratify_cols, sort=False, dropna=False)

        # Global pointer keeps distribution balanced across all labels.
        pointer = 0

        for _, sub_df in grouped:
            sub_df = sub_df.reset_index(drop=True)

            for _, row in sub_df.iterrows():
                group_id = pointer % n_groups
                groups[group_id].append(row)
                pointer += 1

    else:
        # Simple round-robin split without stratification.
        for i, (_, row) in enumerate(df.iterrows()):
            group_id = i % n_groups
            groups[group_id].append(row)

    result = []

    for group_rows in groups:
        if len(group_rows) == 0:
            result.append(pd.DataFrame(columns=df.columns))
        else:
            result.append(
                pd.DataFrame(group_rows, columns=df.columns).reset_index(drop=True)
            )

    return result


# ---------------------------------------------------------------------
# Helper 3: robust Excel header lookup
# ---------------------------------------------------------------------

def find_column_index(header, target_name):
    """
    Find a column index in an Excel header row.

    Returns
    -------
    int
        1-based Excel column index.

    This function is more robust than:
        header.index("label_correctness")

    because it cleans hidden BOM characters and surrounding spaces.
    """

    cleaned_header = [
        str(h).replace("\ufeff", "").strip() if h is not None else ""
        for h in header
    ]

    if target_name not in cleaned_header:
        raise ValueError(
            f"Column '{target_name}' not found.\n"
            f"Available columns are:\n{cleaned_header}"
        )

    return cleaned_header.index(target_name) + 1


# ---------------------------------------------------------------------
# Helper 4: add dropdown menus and formatting to Excel
# ---------------------------------------------------------------------

def add_dropdowns_and_formatting(xlsx_path):
    """
    Add dropdown menus and formatting to one Excel annotation file.

    Dropdown columns:
        label_correctness:
            Correct / Incorrect / Unsure

        missing_edge:
            Yes / No

        missing_node:
            Yes / No

        missing_property_or_qualifier:
            Yes / No

    Free-text column:
        note

    Why multiple Wikidata cause columns?
        A case may involve more than one structural KG issue.
        For example, Wikidata may miss both an entity and a connecting statement.
        Separate Yes/No columns allow annotators to mark all causes that apply.
    """

    wb = load_workbook(xlsx_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    max_row = ws.max_row
    max_col = ws.max_column

    # Find target annotation columns.
    label_correctness_col = find_column_index(header, "label_correctness")
    missing_edge_col = find_column_index(header, "missing_edge")
    missing_node_col = find_column_index(header, "missing_node")
    missing_property_col = find_column_index(header, "missing_property_or_qualifier")

    # Convert column numbers to Excel letters.
    label_correctness_col_letter = get_column_letter(label_correctness_col)
    missing_edge_col_letter = get_column_letter(missing_edge_col)
    missing_node_col_letter = get_column_letter(missing_node_col)
    missing_property_col_letter = get_column_letter(missing_property_col)

    # -----------------------------------------------------------------
    # Dropdown validation
    # -----------------------------------------------------------------

    # Dropdown for taxonomy label correctness.
    dv_label_correctness = DataValidation(
        type="list",
        formula1=f'"{",".join(TAXONOMY_LABEL_CORRECT_OPTIONS)}"',
        allow_blank=True
    )

    # Dropdown for each Wikidata structural cause column.
    dv_yes_no = DataValidation(
        type="list",
        formula1=f'"{",".join(YES_NO_OPTIONS)}"',
        allow_blank=True
    )

    ws.add_data_validation(dv_label_correctness)
    ws.add_data_validation(dv_yes_no)

    if max_row >= 2:
        # Apply label correctness dropdown.
        dv_label_correctness.add(
            f"{label_correctness_col_letter}2:"
            f"{label_correctness_col_letter}{max_row}"
        )

        # Apply Yes/No dropdowns to all three Wikidata cause columns.
        dv_yes_no.add(
            f"{missing_edge_col_letter}2:"
            f"{missing_edge_col_letter}{max_row}"
        )

        dv_yes_no.add(
            f"{missing_node_col_letter}2:"
            f"{missing_node_col_letter}{max_row}"
        )

        dv_yes_no.add(
            f"{missing_property_col_letter}2:"
            f"{missing_property_col_letter}{max_row}"
        )

    # -----------------------------------------------------------------
    # Formatting colors
    # -----------------------------------------------------------------

    header_fill = PatternFill("solid", fgColor="1F4E79")       # dark blue
    header_font = Font(bold=True, color="FFFFFF")              # white text

    row_fill_odd = PatternFill("solid", fgColor="FFFFFF")      # white
    row_fill_even = PatternFill("solid", fgColor="EAF3F8")     # very light blue

    annotation_fill_odd = PatternFill("solid", fgColor="FFF2CC")   # light yellow
    annotation_fill_even = PatternFill("solid", fgColor="FCE4B2")  # deeper yellow

    thin_side = Side(style="thin", color="D9D9D9")

    border = Border(
        top=thin_side,
        left=thin_side,
        right=thin_side,
        bottom=thin_side
    )

    # -----------------------------------------------------------------
    # Header formatting
    # -----------------------------------------------------------------

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    ws.row_dimensions[1].height = 32

    # Clean header for formatting lookup.
    cleaned_header = [
        str(h).replace("\ufeff", "").strip() if h is not None else ""
        for h in header
    ]

    # These are the columns annotators fill in.
    annotation_col_names = [
        "Corrected SPARQL",
        "label_correctness",
        "missing_edge",
        "missing_node",
        "missing_property_or_qualifier",
        "note"
    ]

    annotation_col_indices = []

    for col_name in annotation_col_names:
        if col_name in cleaned_header:
            annotation_col_indices.append(cleaned_header.index(col_name) + 1)

    # -----------------------------------------------------------------
    # Body formatting with alternating row colors
    # -----------------------------------------------------------------

    for row_idx in range(2, max_row + 1):
        is_even_body_row = row_idx % 2 == 0

        normal_fill = row_fill_even if is_even_body_row else row_fill_odd
        annotation_fill = annotation_fill_even if is_even_body_row else annotation_fill_odd

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Annotation columns remain yellow, but still alternate slightly.
            if col_idx in annotation_col_indices:
                cell.fill = annotation_fill
            else:
                cell.fill = normal_fill

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )
            cell.border = border

    # -----------------------------------------------------------------
    # Highlight annotation headers
    # -----------------------------------------------------------------

    for col_idx in annotation_col_indices:
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill("solid", fgColor="FFD966")     # yellow header
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    # -----------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------

    widths = {
        "case_id": 10,
        "question": 45,
        "gold_answer": 35,
        "KG answer": 35,
        "SPARQL": 55,
        "Corrected SPARQL": 55,
        "taxonomy_label": 30,
        "source": 24,
        "label_correctness": 24,
        "missing_edge": 18,
        "missing_node": 18,
        "missing_property_or_qualifier": 30,
        "note": 45,
    }

    for i, col_name in enumerate(cleaned_header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 22)

    # -----------------------------------------------------------------
    # Row heights
    # -----------------------------------------------------------------

    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 65

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Enable Excel filter.
    ws.auto_filter.ref = ws.dimensions

    # Save changes.
    wb.save(xlsx_path)


# ---------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------

def generate_annotation_files(
    input_csv,
    output_dir="annotation_outputs",
    n_groups=3,
    stratify_cols=None
):
    """
    Generate balanced annotation Excel files from one input CSV.

    Parameters
    ----------
    input_csv : str
        Path to the input CSV.

    output_dir : str
        Folder where output Excel files will be saved.

    n_groups : int
        Number of annotation files to create.

    stratify_cols : list[str] or None
        Recommended:
            ["taxonomy_label"]

        This keeps taxonomy label distribution more balanced across groups.

        Use None if you only want simple round-robin balancing.
    """

    input_csv = Path(input_csv)
    output_dir = Path(output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    # Read CSV.
    #
    # header=0 means:
    #     The first row is the column title/header row.
    #     It is NOT treated as data.
    df = pd.read_csv(input_csv, header=0)

    # Clean column names.
    df = clean_column_names(df)

    # Expected input columns.
    expected_cols = [
        "question",
        "gold_answer",
        "KG answer",
        "taxonomy_label",
        "sparql",
        "confidence",
        "source"
    ]

    # Validate input columns.
    missing = [col for col in expected_cols if col not in df.columns]

    if missing:
        raise ValueError(
            f"Missing expected columns: {missing}\n"
            f"Actual columns are: {list(df.columns)}"
        )

    # Remove confidence column because the annotation task should not show it.
    df = df.drop(columns=["confidence"])

    # Rename sparql for the public Excel annotation sheet.
    #
    # The input CSV uses lowercase `sparql` so it can be carried directly from
    # all_valid_cases_with_taxonomy.csv in step 1. The Excel sheet uses `SPARQL`
    # to make the query column easier for annotators to scan.
    df = df.rename(columns={"sparql": "SPARQL"})

    # Add stable case_id.
    #
    # This is useful later when merging annotations from different files.
    # It preserves the original row identity.
    if "case_id" not in df.columns:
        df.insert(0, "case_id", range(1, len(df) + 1))

    # -----------------------------------------------------------------
    # Add annotation columns before splitting
    # -----------------------------------------------------------------

    # This column is left empty by default.
    #
    # Annotators only fill it when the extracted SPARQL query is wrong or needs
    # a corrected version.
    df["Corrected SPARQL"] = ""

    # This column checks whether the automatic taxonomy label is correct.
    df["label_correctness"] = ""

    # Wikidata structural cause columns.
    #
    # These replace the old single wikidata_cause column.
    # Annotators can mark multiple causes as Yes.
    #
    # Example:
    #     missing_edge = Yes
    #     missing_node = No
    #     missing_property_or_qualifier = Yes
    #
    # This means both a missing edge and a missing property/qualifier
    # are relevant for the same case.
    df["missing_edge"] = ""
    df["missing_node"] = ""
    df["missing_property_or_qualifier"] = ""

    # Free-text note column.
    df["note"] = ""

    # Final column order in output files.
    final_cols = [
        "case_id",
        "question",
        "gold_answer",
        "KG answer",
        "SPARQL",
        "Corrected SPARQL",
        "taxonomy_label",
        "source",
        "label_correctness",
        "missing_edge",
        "missing_node",
        "missing_property_or_qualifier",
        "note"
    ]

    df = df[final_cols]

    print("Final columns before splitting:")
    print(list(df.columns))

    # Split into balanced groups.
    splits = split_balanced(
        df,
        n_groups=n_groups,
        stratify_cols=stratify_cols
    )

    output_paths = []

    # Save each group to an Excel file.
    for i, split_df in enumerate(splits, start=1):
        out_path = output_dir / f"annotation_group_{i}.xlsx"

        print(f"\nSaving group {i}: {len(split_df)} rows")
        print("Columns:", list(split_df.columns))

        # Save dataframe to Excel.
        split_df.to_excel(out_path, index=False)

        # Add dropdowns and formatting.
        add_dropdowns_and_formatting(out_path)

        output_paths.append(out_path)

    # Print summary.
    print("\nFinished.")
    print(f"Total input data rows: {len(df)}")

    for i, split_df in enumerate(splits, start=1):
        print(f"group {i}: {len(split_df)} rows")

    print("\nOutput files:")

    for path in output_paths:
        print(path)

    return output_paths


def main() -> None:
    """Run step 1 and step 2 as one merged annotation pipeline."""
    run_step1_build_annotation_csvs()

    try:
        project_root = find_project_root()
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return

    main_csv = project_root / MODEL_PAIR["main_csv"]
    output_dir = project_root / "Annotation" / "annotation_outputs"

    if not main_csv.exists():
        print(f"ERROR: Step 1 output was not found: {main_csv}")
        return

    print()
    print("=" * 72)
    print("Step 2: Excel annotation split")
    print("=" * 72)

    generate_annotation_files(
        input_csv=main_csv,
        output_dir=output_dir,
        n_groups=3,

        # Recommended for your task:
        # balance the distribution of taxonomy labels across the 3 files.
        #
        # Do NOT stratify by source because source may be unique per row.
        stratify_cols=["taxonomy_label"]

        # Alternative:
        # Use this if you want pure round-robin splitting:
        # stratify_cols=None
    )


if __name__ == "__main__":
    main()
